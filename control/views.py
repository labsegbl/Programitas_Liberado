# Imports estándar de Python
import csv, io , json, os, re, threading
from collections import defaultdict
from datetime import datetime
from time import sleep
from urllib.parse import quote, unquote_plus
import ipaddress
import pyotp
import qrcode
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

# Imports de terceros
import requests
import pandas as pd
from dateutil.relativedelta import relativedelta
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select, WebDriverWait
from decouple import config

# Imports de Django
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.views import LoginView
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import IntegrityError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.utils import timezone
from django.utils.http import urlencode
from django.views.generic import View, TemplateView

# Imports locales
from .forms import *
from .models import *
from .resources import *


#Redireccionamiento a las platillas de html que se crearon 
def index(request , mensaje=None): #Plantilla de logeo para los usuarios 
    return render(request, 'login.html', {'error':mensaje})

@login_required 
def inicio(request): #Plantilla de la primera pagina al entrar al sistema 
    return render(request , 'inicio.html' , {'usuario':request.user} )

#=============================================================================================
#        Funciones para el control de acceso de usuarios 
#=============================================================================================

@login_required
def crearUsuario(request): #Se toman los datos de la plantilla de registrarse para asi tomar los campos y crear un nuevo usuario con el sistema de autenticación de Django
    if request.method == 'GET':
        return render(request , 'registrarse.html') #La plantilla donde se establece el formulario para ingresar los datos
    else: #Obtenemos los datos del nuevo usuario
        if request.POST['password'] == request.POST['password1']: #Se comprueba que la contraseña y la verificación sean las mismas
            username = request.POST.get('username')
            nombres = request.POST.get('name')
            apellidos = request.POST.get('lastname')
            email = request.POST.get('email')
            password = request.POST.get('password')

            user = User.objects.create_user(username, email, password) # Se crea un usuario en base al model User propio de Django
            user.first_name = nombres #Añadimos datos personales al usuario creado
            user.last_name = apellidos
            user.save() # Almacenamos el usuario en la DB 
            return render(request , 'registrarse.html', {'exito':"Usuario creado correctamente"})
        else :
            return render(request , 'registrarse.html', {'advertencia':"Las contraseñas no coinciden"})

def validarUsuario(request): #Función para validar los usuarios al momento de logearse 
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        otp_code = request.POST.get('otp_code' , None) # Codigo de 2FA en caso de existir 
        user = authenticate(request, username=username, password=password) # Verificamos al usuario y lo autenticamos si las credenciales son correctas
        
        if user is not None:
            if user.otp_secret:
                totp = pyotp.TOTP(user.otp_secret) # Encuentra el usuario y busca el código OTP correspondiente y analiza si es correcto
                if not totp.verify(otp_code):
                    return render(request, 'login.html' , {'error': "El código OTP es incorrecto."})

            login(request, user) # Se logeea el usuario, este hace uso de cookies para navegar entre las paginas
            return redirect('detector') # Se redirecciona a la pagina principal
        else:
            return render(request, 'login.html' , {'error': "Algó salio mal en la autenticación \nCompruebe que su usuario y contraseña sean correctos"})
    else:
        return render(request, 'login.html')

def signout(request): # Función para cerrar seción y se borre las cookies de credenciales del navegador
    logout(request=request)
    return redirect('index')

def delete_file(file_path): # Función que borra las imagenes de los codigos QR de OTP
    if os.path.isfile(file_path):
        os.remove(file_path)

@login_required
def setup_2fa(request): # Fución para activar y generar el 2FA con OTP, muestra el código en pantalla
    user = request.user
    if not user.otp_secret:   
        totp = pyotp.TOTP(pyotp.random_base32()) # Genera un nuevo código secreto si este no existe
        user.otp_secret = totp.secret # Lo guarda en la variable de usuario otp_secret
        user.save()
    
    totp = pyotp.TOTP(user.otp_secret) # Genera el código QR
    qr = qrcode.QRCode( # Genera la imagen del QR y se establecen los parametros para que se integren en la imagen.
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    uri = totp.provisioning_uri(name=user.email, issuer_name='ProgramitasSI')
    qr.add_data(uri)
    qr.make(fit=True)
    img = qr.make_image(fill='black', back_color='white')
    
    buffer = BytesIO() # Guardar la imagen en un archivo temporal
    img.save(buffer, format='PNG')
    buffer.seek(0)
    
    file_name = f'qr_code_{user.id}.png' # Guarda la imagen en un archivo de la carpeta de media 
    file_path = os.path.join(settings.MEDIA_ROOT, file_name)
    
    if not os.path.exists(settings.MEDIA_ROOT):
        os.makedirs(settings.MEDIA_ROOT)
    
    with open(file_path, 'wb') as f: # Escribe la imagen en el directorio
        f.write(buffer.getvalue())
    
    # URL del código QR
    qr_code_url = os.path.join(settings.MEDIA_URL, file_name)
      
    def delete_file_task():# Eliminar la imagen de QR después de 30 segundos
        import time
        time.sleep(30)
        delete_file(file_path)
    
    threading.Thread(target=delete_file_task).start() # Se ejecuta la operación en paralelo 
    
    context = {'qr_code_url': qr_code_url}
    return render(request, 'setup_2fa.html', context)

#=============================================================================================
#                      Verificaciones de validación de IPs
#=============================================================================================

regex = "^((25[0-5]|2[0-4][0-9]|1[0-9][0-9]|[1-9]?[0-9])\.){3}(25[0-5]|2[0-4][0-9]|1[0-9][0-9]|[1-9]?[0-9])$" # Formula regex para validar que una IP sea valida

@login_required
def validarIPValida(request, ip):  # Función para determinar si una ip es valida o no 
    return True if(re.search(regex, ip)) else False

@login_required
def ipEsMayor(request, mayor, menor): # Función para determinar si una ip es mayor a otra 
    ip1 = ipaddress.ip_address(mayor)
    ip2 = ipaddress.ip_address(menor)

    try:  # Comparar las direcciones IP para verificar que el rango sea valido
        if ip1 > ip2:
            return True
        elif ip1 < ip2:
            return False
        else:
            return False
    except ValueError as e:
        return f"Error en la dirección IP: {e}" 

@login_required
def ingresarRangoIps(request): # Función para exonerar rangos de IPs 
    if request.method == 'POST':
        ips = IP.objects.all()
        ipInicio = str(request.POST['ipInicio']).strip()#Toma las IPs ingresadas
        ipFin = str(request.POST['ipFin']).strip()

        if validarIPValida(request, ipInicio) and validarIPValida(request, ipFin): # Valida las IPs para determinar si son validas o no y si lo son las ingresa en un rango exonerado
            if (ipEsMayor(request,ipFin, ipInicio)):
                RangoExonerado.objects.create(ipInicio=ipInicio ,ipFin=ipFin)
            else:
                RangoExonerado.objects.create(ipInicio=ipFin,ipFin=ipInicio)
                messages.error(request, "El rango de IPs se ingreso con exito")
            return redirect('detector')
        else:     
            messages.error(request, "Las IPs ingresadas no son validas")  
            return redirect('detector')

def obtenerDatosDeAbuse(ip): #Función para traer los datos del API de AbuseIpDB
    # Defining the api-endpoint
    url = 'https://api.abuseipdb.com/api/v2/check'    
    querystring = {'ipAddress': ip}   
    headers = {
        'Accept': 'application/json',
        'Key': config("ABUSE_APIKEY")#apikey de la cuenta de Georky
    }
    
    try:
        response = requests.request(method='GET', url=url, headers=headers, params=querystring)
        print(json.loads(response.text))
    except Exception as e:
        print("Error al traer datos del API de abuse \n",e)

    return json.loads(response.text) #Se devuelve la respuesta en formato de JSON para que manipular los datos sea más fácil

#=============================================================================================
#                Control de estado de IPs
#=============================================================================================
def bloquearCambios(): #Cambios realizados en el DBMS
    ips = IP.objects.all()
    for ip in ips:
        if ip.estado == 'Bloqueado': #verifica que si se bloquearon ips en el administrador estos cambios se reflejen en el sistema
            objetoBloqueado, creadoBloqueado = Historial_IP_FW_Bloqueadas.objects.get_or_create(ipBloqueada=ip)
            if not creadoBloqueado:
                objetoBloqueado.save()

@login_required
def mandarACasosEspeciales(request, ip , caso): #Función para controlar casos especiales al tratar las IPs
    ipPermitida = Historial_IP_FW_Permitidas.objects.filter(ipPermitida=ip)
    ipBloqueada = Historial_IP_FW_Bloqueadas.objects.filter(ipBloqueada=ip)

    ipPermitida.delete() if ipPermitida.exists() else None
    ipBloqueada.delete() if ipBloqueada.exists() else None # Si es especial se elimina de lista blanca o negra 

    if caso == 'ipDeEcuador': # En caso de que la IP sea de Ecuador se manda directo a pendiente con ese motivo
        ip.estado = "Pendiente"
        ip.save()
        ipEspeciales = Casos_Especiales(ipEspecial = ip, razon="Ip de Ecuador")
        ipEspeciales.save() #Creamos un caso especial con esta direccion IP

@login_required
def ingresarNueva(request , ip , fuente , opcion='option2' , row=None):
    try:
        decodedResponse = obtenerDatosDeAbuse(ip) #Traemos los datos de abuse
        pais = decodedResponse["data"]["countryCode"] # Formateamos los datos en base a las respuestas traidas por el Abuse
        malicioso = int(decodedResponse["data"]["abuseConfidenceScore"])
        dominio = decodedResponse["data"]["domain"]
        isp = decodedResponse["data"]["isp"]
        tipoUso = decodedResponse["data"]["usageType"]

        if decodedResponse["data"]["isPublic"] == True: # Con esta verificación revisamos si la IP es publica o privada
            objetoPais, creadoPais = Pais.objects.get_or_create(iso2=pais) # Traemos o creamos el país de origen de la IP ya que debe existir segun el modelo de la DB
            objectoDominio, creadoDominio = Dominio.objects.get_or_create(nomDominio=dominio)# Traemos o creamos el dominio de la IP
            

            if IP.objects.filter(ip=ip).exists(): #Verificamos si ya existe la direccion IP. Si existe controlamos el caso de si esta en la lista blanca
                actual = IP.objects.filter(ip=ip)[0]

                if not Historial_IP_FW_Bloqueadas.objects.filter(ipBloqueada=actual).exists():
                    if Historial_IP_FW_Permitidas.objects.filter(ipPermitida=actual).exists():
                        hPermitida = Historial_IP_FW_Permitidas.objects.filter(ipPermitida=actual)
                        hPermitida.delete()

                        ipEspeciales = Casos_Especiales(ipEspecial = actual)
                        ipEspeciales.save() #Creamos un caso especial con esta direccion IP
                    else:
                        actual.estado = "Pendiente"
                        actual.save() #Actualizamos los datos del registro de la IP 
                else:
                    print("Ip que ya esta bloqueda")

                mandarACasosEspeciales(request, actual, 'ipDeEcuador') if objetoPais.nombre == 'Ecuador' else None

            else: # Se crea la nueva direccion IP       
                if opcion == 'option1': # En caso que la entrada del excel tenga formato
                    nuevaIP = IP(ip=ip , estado="Bloqueado", malicioso=malicioso, isp=isp,
                                tipoUso=tipoUso , pais=objetoPais, dominio=objectoDominio,
                                ataques=row[3], descripcion=row[4], peticiones=int(row[5]),
                                firewall="NO", usuario=request.user.username, fuente=fuente)
                else : # Si el escel de ingreso solo tiene IPs
                    nuevaIP = IP(ip=ip , estado="Bloqueado", malicioso=malicioso, isp=isp,
                                tipoUso=tipoUso , pais=objetoPais, dominio=objectoDominio,
                                firewall="NO", usuario=request.user.username, fuente=fuente)
                nuevaIP.save()

                if DominioPermitido.objects.filter(nomDominio=objectoDominio).exists(): #Si la Ip se encuentra en un dominio permitido se asigna un estado de No Bloqueo
                    print("Ingreso al dominio permitido")
                    nuevaIP.estado = "No Bloqueado"
                    nuevaIP.save()
                    ipEspeciales = Casos_Especiales(ipEspecial = nuevaIP , razon = "Ip con Dominio Exonerado")
                    ipEspeciales.save() #Creamos un caso especial con esta direccion IP
                elif seEncuentraEnRango(ip): # Si se encuentra la IP en un rango exonerado se le asigna un estado de No Bloqueado
                    print("Ingreso en un rango exonerado")
                    nuevaIP.estado = "No Bloqueado"
                    nuevaIP.save()
                    ipEspeciales = Casos_Especiales(ipEspecial = nuevaIP , razon = "Ip con Rango Exonerado")
                    ipEspeciales.save() #Creamos un caso especial con esta direccion IP
                else:    
                    print("Ingreso al Bloqueo")    
                    nuevaBloqueada = Historial_IP_FW_Bloqueadas(ipBloqueada=nuevaIP)
                    nuevaBloqueada.save() #Creamos la nueva IP y se establece como bloqueada

                mandarACasosEspeciales(request, nuevaIP, 'ipDeEcuador') if objetoPais.nombre == 'Ecuador' else None # En caso de la IP ser de Ecuador se la exonera
        else: # Si la IP es privada se le asignan valores quemados referentes a indicar que es una IP Privada
            objetoPais = Pais.objects.get(iso3='000')
            objetoDominio = Dominio.objects.get(nomDominio='Red Privada')
            nuevaIP = IP(ip=ip , estado="No Bloqueado", malicioso=malicioso, isp=isp,
                                tipoUso="Desconocido" , pais=objetoPais, dominio=objetoDominio,
                                firewall="NO", usuario=request.user.username, fuente=fuente)
            ipEspeciales = Casos_Especiales(ipEspecial = nuevaIP , razon = "Ip Privada")
            nuevaIP.save()
            ipEspeciales.save()
            
    except Exception as e:
        print(e , "al ingresar")
        return ip

@login_required
def establecerBloqueoTemporal(request, ip, fechaInicio, fechaFin): # Función para tratar con las IPs con Estado temporal establecido

    fechaInicio = datetime.strptime(fechaInicio, '%Y-%m-%d').date() if fechaInicio else None # Tomamos las fechas entre las cuales se ejecutara el bloqueo
    fechaFin = datetime.strptime(fechaFin, '%Y-%m-%d').date() if fechaFin else None

    if fechaFin == None and fechaInicio != None: # Si hay fecha de inicio pero no de fin se establece esta ultima un año más
        fechaFin = fechaInicio + relativedelta(years=1)
    elif fechaInicio == None and fechaFin != None: # Si no hay fecha de inicio peor si de fin, la fecha de inicio es hoy
        fechaInicio = datetime.today().date()
    if fechaInicio > fechaFin : # Se analiza que la fecha de inicio sea menor a la de fin 
        aux = fechaInicio
        fechaInicio = fechaFin
        fechaFin = aux 


    objetoBloqueada, creadoBloqueada = Historial_IP_FW_Bloqueadas.objects.get_or_create(ipBloqueada = ip) # Traemos la IP bloqueada o la creamos en caso de no existir 
    nueva = BloqueadasTemporales(ipBloqueada=objetoBloqueada , fechaInicio=fechaInicio if fechaInicio else None , fechaFin=fechaFin) # Creamos un nuevo Bloqueo Temporal 
    nueva.save()

    # Organizamos para que el bloqueo se ejecute entre las fechas ingresadas
    if fechaInicio <= datetime.now().date() <= fechaFin:
        ip.estado = "Bloqueado Temporal"
    else:
        ip.estado = "Pendiente" # Estado alterno de cuando la IP con bloqueo temporal no se encuentra en su periodo de bloqueo establecido
    ip.save()

    if fechaFin < datetime.now().date(): # En caso de que el bloqueo haya salido de temporal se manda un caso especial con esa ip 
        objetoBloqueada.delete()
        finTemporal = Casos_Especiales(ipEspecial=ip , razon="Ip salida de Temporal")
        finTemporal.save()

@login_required
def revisarIPsTemporales(request): # Revisa los periodos de bloqueo de cada una de las IPs establecidas con bloqueo temporal
    ipsTemporales = BloqueadasTemporales.objects.all()
    
    for ip in ipsTemporales: # Si la IP esta entre las fechas de bloqueo se mantiene
        if ip.fechaInicio.date() <= datetime.now().date() <= ip.fechaFin.date() and ip.ipBloqueada.ipBloqueada.estado != "Bloqueado Temporal":
            inicio = IP.objects.filter(ip=ip.ipBloqueada.ipBloqueada.ip)[0]
            inicio.estado = "Bloqueado Temporal"
            inicio.save()
        if ip.fechaFin.date() < datetime.now().date(): # Caso contrario se crea un caso especial con la IP 
            razon = "Ip salida de Temporal el " + str(ip.fechaFin.date)
            finTemporal = Casos_Especiales(ipEspecial=ip.ipBloqueada.ipBloqueada , razon=razon)
            ip.ipBloqueada.ipBloqueada.estado = "Pendiente"
            ip.ipBloqueada.ipBloqueada.save()
            finTemporal.save()
            ip.ipBloqueada.delete()

@login_required
def anadirIndividual(request): # Función para el ingreso de IPs de forma manual, se ingresa una a la vez
    if request.method == 'POST':
        try:
            ip = str(request.POST['ipIndividual']).strip() # Se toma la dirección IP
            fuente = request.POST['fuente']
            verificacion = ingresarNueva(request , ip, fuente) # Se manda a ingresar la IP y que esta siga toda la verificación en la función 

            if request.POST.get('filtros_agregados') == 'on': #Agregamos la función de bloqueo temporal en caso de estar activado el parametro 
                ipPermitida = Historial_IP_FW_Permitidas.objects.filter(ipPermitida=IP.objects.filter(ip=ip)[0])
                ipPermitida.delete() if ipPermitida.exists() else None 
                establecerBloqueoTemporal(request, IP.objects.filter(ip=ip)[0] , request.POST.get('fechaInicio') , request.POST.get('fechaFin') )

            messages.error(request, "Falla en ingreso de la IP") if verificacion else messages.success(request, "Ip ingresada con exito.")
        except Exception as e:
            IP.objects.filter(ip = ip).delete()
            messages.error(request, "Falla en ingreso de la IP ,"+str(e))
            pass
        return redirect('detector')
        
@login_required
def anadirANegra(request , id , rutaRetorno):
    try:
        ip = IP.objects.get(pk=id) #Buscamos la direccion IP en la DB 
        ip.estado = "Bloqueado"
        ip.save() #Cambiamos el estado de la IP
        nuevaIP = Historial_IP_FW_Bloqueadas(ipBloqueada=ip) #Creamos la IP en la lista negra 
        nuevaIP.save()

        objetoPermitido = Historial_IP_FW_Permitidas.objects.filter(ipPermitida=ip.id) #En caso que esta haya estado en la lista blanca se la retira 
        objetoPermitido.delete() if objetoPermitido else None

        objetoRevision = Casos_Especiales.objects.filter(ipEspecial=ip.id)
        objetoRevision.delete() if objetoRevision else None
    except:
        messages.error(request, "Direccion IP ya existente en bloqueadas.")
        return redirect('detector')
    return redirect(str(unquote_plus(rutaRetorno))[12:]) #Retorna a la pagina anterior

@login_required
def anadirABlanca(request , id , rutaRetorno ):
    try:
        ip = IP.objects.get(pk=id) #Buscamos la direccion IP en la DB 
        ip.estado = "Exonerado"
        ip.save()#Cambiamos el estado de la IP
        nuevaIP = Historial_IP_FW_Permitidas(ipPermitida=ip) #Creamos la IP en la lista blanca 
               
        objetoBloqueado = Historial_IP_FW_Bloqueadas.objects.filter(ipBloqueada=ip.id)#En caso que esta haya estado en la lista negra se la retira 
        objetoBloqueado.delete() if objetoBloqueado else None

        objetoRevision = Casos_Especiales.objects.filter(ipEspecial=ip.id)
        if objetoRevision:
            nuevaIP.descripcion = objetoRevision.first().razon
            objetoRevision.delete() 
        else: None
        nuevaIP.save()
    except:
        messages.error(request, "Direccion IP ya existente en permitidas.")
        return redirect('detector')
    return redirect(str(unquote_plus(rutaRetorno))[12:]) #Retorna a la pagina anterior

def seEncuentraEnRango( ip): # Función para validar si una IP se encuentra en un rango exonerado
    ipVal = ipaddress.ip_address(ip)
    rangos = RangoExonerado.objects.all()

    for rango in rangos: # Analiza si la IP se encuentra en alguno de los rangos almacenados
        ipIni = ipaddress.ip_address(rango.ipInicio)
        ipFin = ipaddress.ip_address(rango.ipFin)

        if (ipIni<=ipVal and ipVal<=ipFin):
            return True
    
    return False
       
#=============================================================================================
#                              Manejo de IPs
#=============================================================================================

@login_required
def listaNegra(request):
    revisarIPsTemporales(request)
    ips = Historial_IP_FW_Bloqueadas.objects.all()
    if request.method == 'POST':
        archivo_subido = request.FILES['archivoEntrada'] #Se obtiene el archivo con nuevas IPs a bloquear
        if archivo_subido.name.endswith('.xlsx'):
            filas = iter(openpyxl.load_workbook(archivo_subido).active.iter_rows(values_only=True)) #Tomamos las filas del excel 

            for fila in filas:
                i = obtenerDatosDeAbuse(fila[0])["data"] #Sacamos la data de los datos traidos por el Abuse
                objetoPais, creadoPais = Pais.objects.get_or_create(iso2=i["countryCode"]) # Creamos o traemos el pais 
                objectoDominio, creadoDominio = Dominio.objects.get_or_create(nomDominio=i["domain"]) # Creamos o traemos el dominio
                
                objetoIp, creadoIp = IP.objects.get_or_create(ip=fila[0] 
                                                              ,defaults= {'malicioso':int(i["abuseConfidenceScore"]), 'pais':objetoPais, 'estado': "Bloqueado"
                                                                          ,'dominio':objectoDominio, 'isp':i["isp"], 'tipoUso':i["usageType"] ,'usuario':request.user})
                objetoIpBloqueada, creadoIpBloqueada = Historial_IP_FW_Bloqueadas.objects.get_or_create(ipBloqueada = objetoIp)
                objetoIpBloqueada.save()   #Creamos la nueva IP y la mandamos a bloquear en lista negra 
            
            ips = Historial_IP_FW_Bloqueadas.objects.all()
            return render(request , 'ipBloqueadas.html', {'ips': ips , 'rutaRetorno': urlencode({'rutaRetorno': request.path})}) #Actualizamos las IPs y volvemos al panel de visualización
        else:
            return render(request , 'ipBloqueadas.html', {'ips': ips , 'advertencia': 'El archivo debe tener formato .xlsx' , 'rutaRetorno': urlencode({'rutaRetorno': request.path})})
    else:
        return render(request , 'ipBloqueadas.html', {'ips': ips , 'rutaRetorno': urlencode({'rutaRetorno': request.path})})

@login_required
def listaBlanca(request):
    ips = Historial_IP_FW_Permitidas.objects.all()
    if request.method == 'POST':
        archivo_subido = request.FILES['archivoEntrada'] #Se obtiene el archivo con nuevas IPs a exonerar
        if archivo_subido.name.endswith('.xlsx'):
            filas = iter(openpyxl.load_workbook(archivo_subido).active.iter_rows(values_only=True)) #Tomamos las filas del excel 

            for fila in filas:
                i = obtenerDatosDeAbuse(fila[0])["data"] #Sacamos la data de los datos traidos por el Abuse
                objetoPais, creadoPais = Pais.objects.get_or_create(iso2=i["countryCode"]) # Creamos o traemos el pais 
                objectoDominio, creadoDominio = Dominio.objects.get_or_create(nomDominio=i["domain"]) # Creamos o traemos el dominio
                
                objetoIp, creadoIp = IP.objects.get_or_create(ip=fila[0] 
                                                              ,defaults= {'malicioso':int(i["abuseConfidenceScore"]), 'pais':objetoPais, 'estado': "Exonerado"
                                                                          ,'dominio':objectoDominio, 'isp':i["isp"], 'tipoUso':i["usageType"] ,'usuario':request.user})
                objetoIpPermitida, creadoIpPermitida = Historial_IP_FW_Permitidas.objects.get_or_create(ipPermitida = objetoIp)
                objetoIpPermitida.descripcion = request.POST['descripcion']
                objetoIpPermitida.save()    #Creamos la nueva IP y la mandamos a exonerar en lista blanca 
            
            ips = Historial_IP_FW_Permitidas.objects.all()
            return render(request , 'ipPermitidas.html', {'ips': ips , 'rutaRetorno': urlencode({'rutaRetorno': request.path})}) #Actualizamos las IPs y volvemos al panel de visualización
        else:
            return render(request , 'ipPermitidas.html', {'ips': ips , 'advertencia': 'El archivo debe tener formato .xlsx' , 'rutaRetorno': urlencode({'rutaRetorno': request.path})})
    else:
        return render(request , 'ipPermitidas.html', {'ips': ips , 'rutaRetorno': urlencode({'rutaRetorno': request.path})})

@login_required
def leer_archivo(request, archivo_subido, tipo_entrada, fuente): #Función para tratar los archivos subidos 
    ipsConError = [] # Una lista para tomar las ips que no se pudieron ingresar e informar de ello 
    if archivo_subido.name.endswith('.xlsx'): # En caso de ser archivo de excel 
        try:
            wb = openpyxl.load_workbook(archivo_subido) 
            sheet = wb.active #Tomamos la primera hoja del archivo 
            filas = sheet.iter_rows(values_only=True) #Extraemos las filas  
            next(filas) if tipo_entrada == 'option1' else None # La opcion 1 hace referencia a un archivo con formato 
            for row in filas:
                try:
                    ip = str(row[0]).strip().replace(',','.') # Tomamos el campo de IP y lo mandamos a ingresar a la DB
                    print(ip)
                    ingreso = ingresarNueva(request, ip, fuente, tipo_entrada, row) # Sigue el proceso para verificar el estado que asume 
                except Exception as e:
                    print(f'Error al analizar IP: {str(e)}')

                ipsConError.append(ingreso) if ingreso else None # En caso de existir error al ingresar la almacenamos 
        except Exception as e:
            print(f'Error al procesar el archivo Excel: {str(e)}')
    elif archivo_subido.name.endswith('.txt'): # En caso de ser un archivo txt 
        try:
            for linea in archivo_subido:
                ip = linea.decode('utf-8').strip().replace(',','.') # Decodificamos el archivo con utf-8 para evitar el cambio o aumento de caracteres 
                print(ip) 
                try:
                    ingreso = ingresarNueva(request, ip, fuente) # Mandamos a ingresar la IP
                except:
                    pass

                ipsConError.append(ingreso) if ingreso else None # En caso de existir error al ingresar la almacenamos 
        except Exception as e:
            print(f'Error al procesar el archivo de texto: {str(e)}')

    return ipsConError #Devolvemos todas las IPs con error para presentarlas al usuario 

@login_required
def detector(request): # Pagina render del visualizador de IPs 
    ips = IP.objects.all()

    if request.method == 'POST':
        archivo_subido = request.FILES['archivoEntrada'] # Tomamos el archivo de IPs 
        tipo_entrada = request.POST['tipoEntrada']
        fuente = request.POST['fuente']

        if archivo_subido.name.endswith('.xlsx') or archivo_subido.name.endswith('.txt'): # Verificamos que el archivo sea excel o txt caso contrario no lo procesara 
            ipsConError = leer_archivo(request, archivo_subido, tipo_entrada , fuente) # Mandamos a leer el archivo y traemos las IPs que no se pudieron procesar 
            ips = IP.objects.all() # Actualizamos las IPs de casos especiales para que el usuario pueda asignarles un estado 

            if len(ipsConError)> 0: # Organizamos el mensaje final de usuario en caso de que no se puedan procesar las IPs 
                mensajeNoSubidas = "Error al procesar la(s) ip(s): "
                for ip in ipsConError:
                    mensajeNoSubidas += str(ip)+"     ,     "
                mensajeNoSubidas += "  recuerde revisar las IPs pendientes." # Mandamos el mensaje con las IPs que no se pudieron procesar 
                return render(request , 'detector.html' , {'ips': ips , 'rutaRetorno': urlencode({'rutaRetorno': request.path}), 'error': mensajeNoSubidas})
            else: # Si todas las IPs son correctas mandamos un mensaje de exito al ingresar 
                return render(request , 'detector.html' , {'ips': ips , 'rutaRetorno': urlencode({'rutaRetorno': request.path}), 'exito': 'Archivo cargado con éxito.'})
        else: # Si da error se debe devolver a la pagina principal de detección de IPs 
            return render(request , 'detector.html' , {'ips': ips , 'rutaRetorno': urlencode({'rutaRetorno': request.path}), 'advertencia': 'El archivo debe tener formato .xlsx o .txt'})
    else:
        return render(request , 'detector.html' , {'ips': ips , 'rutaRetorno': urlencode({'rutaRetorno': request.path})})

@login_required
def enRevision(request): # Para llamar al final de la carga de archivos los casos especiales o pendientes de IPs
    revisarIPsTemporales(request)
    especiales = Casos_Especiales.objects.all()
    return render(request , 'ipRevision.html' , {'especiales':especiales , 'rutaRetorno': urlencode({'rutaRetorno': request.path})})

@login_required
def anadirDominioExonerado(request): # Función para tratar IPs especificas que provienen de un Dominio Especifico
    ips = Historial_IP_FW_Permitidas.objects.all()
    if request.method == 'POST':
        try:
            objectoDominio, creadoDominio = Dominio.objects.get_or_create(nomDominio=request.POST['dominioExonerado']) # Creamos o traemos el dominio

            nuevo = DominioPermitido(nomDominio=objectoDominio)
            nuevo.save()
            return render(request , 'ipPermitidas.html', {'ips': ips , 'exito': 'Dominio ingresado con exito' , 'rutaRetorno': urlencode({'rutaRetorno': request.path})})
        except:
            return render(request , 'ipPermitidas.html', {'ips': ips , 'error': 'No se pudo ingresar el Dominio' , 'rutaRetorno': urlencode({'rutaRetorno': request.path})})
        
#=============================================================================================
#                         Manipulación de IPs
#=============================================================================================

@login_required
def eliminar(request , id , rutaRetorno): #Elimina una direccion ip especifica
    ip = IP.objects.get(pk=id)
    ip.delete()
    return redirect(str(unquote_plus(rutaRetorno))[12:]) #Retorna a la pagina anterior

@login_required
def noBloquear(request , id):# Cambia a estado de No Bloquear una IP especifica
    ip = IP.objects.get(pk=id)
    ip.estado = "No Bloqueado"
    ip.save()
    return redirect('enRevision')

@login_required
def enPendiente(request , id): # Cambia a estado de Pendiente una IP especifica
    ip = IP.objects.get(pk=id)
    ip.estado = "Pendiente"
    ip.save()
    return redirect('enRevision')

@login_required
def salirRevision(request , id): # Elimina una Ip especifica de casos especiales
    ip = IP.objects.get(pk=id)
    caso = Casos_Especiales.objects.get(ipEspecial=ip)
    caso.delete()
    return redirect('enRevision')

@login_required
def modificar(request , id , rutaRetorno):
    ip = IP.objects.get(pk=id) #Obtenemos la IP de la cual se quieren modificar los datos
    if request.method == 'POST':
        form = IPForm(request.POST,instance=ip) #Traemos los datos ya existentes de la DB referentes a la IP
        if form.is_valid():
            form.save()# Guardamos los datos y segun el estado mandamos a verificar a las diferentes funciones 
            if form.cleaned_data['estado'] == "Bloqueado":
                anadirANegra(request , id , rutaRetorno)
            elif form.cleaned_data['estado'] == "Exonerado":
                anadirABlanca(request , id , rutaRetorno)
            elif form.cleaned_data['estado'] == "Pendiente"  or form.cleaned_data['estado'] == "No Bloqueado":
                try:
                    caso = Casos_Especiales(ipEspecial = ip)
                    caso.save()

                    objetoPermitido = Historial_IP_FW_Permitidas.objects.filter(ipPermitida=ip.id) #En caso que esta haya estado en la lista blanca se la retira 
                    objetoPermitido.delete() if objetoPermitido else None

                    objetoBloqueado = Historial_IP_FW_Bloqueadas.objects.filter(ipBloqueada=ip.id)#En caso que esta haya estado en la lista negra se la retira 
                    objetoBloqueado.delete() if objetoBloqueado else None

                except:pass
            ips = IP.objects.all() #Obtenemos los datos actualizados para redirigir al registro
            return redirect(str(unquote_plus(rutaRetorno))[12:]) #Retorna a la pagina anterior
        elif form.errors:
            return render(request, 'modificarIP.html', {'form': form , 'advertencia': "Asegurese de llenar todos los campos o que estos sean correctos" })
    else:           
        form = IPForm(instance=ip)
    return render(request, 'modificarIP.html', {'form': form})

@login_required
def exportar(request , tipo): #El tipo indicara que exportar. 1:Todo, 2:Bloqueadas, 3:Exoneradas
    ips_resource = IPResource()
    if(tipo == 1): #Traemos todas las ips para exportar
        ips = IP.objects.all()
        nombre = "datosIP" #Nombre del archvio 
        dataset = ips_resource.export(ips)
    elif(tipo == 2):
        ipsBloqueadas = Historial_IP_FW_Bloqueadas.objects.all() #Traemos la lista negra para exportar 
        ips = [ip.ipBloqueada for ip in ipsBloqueadas]
        nombre = "datosIPsBloqueadas"
        dataset = ips_resource.export(ips)
    elif(tipo == 3):
        ipsPermitidas = Historial_IP_FW_Permitidas.objects.all() #Traemos la lista blanca para exportar 
        ips = [ip.ipPermitida for ip in ipsPermitidas]
        nombre = "datosIPsPermitidas"
        dataset = ips_resource.export(ips)
    else:
        ips = IP.objects.all() #Por defecto y excepcion se exporta todo 
        nombre = "datosIP"
        dataset = ips_resource.export(ips)

    response = HttpResponse(dataset.xlsx, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre}.xlsx"'

    return response

@login_required
def accionesMultiples(request): #Aplicamos acciones de cambios a multiples direcciones IPs
    if request.method == 'POST':
        data = json.loads((request.body).decode('utf-8')) # Se usa request.body ya que se estan enviando los datos desde js y no desde un form 
        ids = data.get('ids', []) # Traemos las direcciones IPs
        accion = data.get('accion', '') # Traemos la accion que se desea realizar con las ips

        if accion == 'eliminar': #Eliminamos las IPs
            for id_ip in ids:
                IP.objects.filter(pk=id_ip).delete()
        elif accion == 'bloquear': #Cambiamos el estabo a Bloqueado a todas las IPs que lleguen 
            for id_ip in ids:
                ip = IP.objects.get(pk=id_ip)
                ip.estado = "Bloqueado"
                ip.save()
                try:
                    nuevaIP = Historial_IP_FW_Bloqueadas(ipBloqueada=ip) #La guardamos en la lista negra 
                    nuevaIP.save()

                    objetoPermitido = Historial_IP_FW_Permitidas.objects.filter(ipPermitida=ip.id) # Eliminamos de lista blanca en caso de existir
                    objetoPermitido.delete() if objetoPermitido else None

                    objetoRevision = Casos_Especiales.objects.filter(ipEspecial=ip.id) #Eliminamos de casos especiales en caso de existir 
                    objetoRevision.delete() if objetoRevision else None
                except:
                    pass
        elif accion == 'exonerar': #Cambiamos el estabo a Exonerado a todas las IPs que lleguen 
            for id_ip in ids:
                ip = IP.objects.get(pk=id_ip)
                ip.estado = "Exonerado"
                ip.save()
                try:
                    nuevaIP = Historial_IP_FW_Permitidas(ipPermitida=ip) #La guardamos en la lista blanca 
                    nuevaIP.save()
                    
                    objetoBloqueado = Historial_IP_FW_Bloqueadas.objects.filter(ipBloqueada=ip.id) # Eliminamos de lista negra en caso de existir
                    objetoBloqueado.delete() if objetoBloqueado else None
                except:
                    pass
        elif accion == 'pendiente':
            for id_ip in ids: #Cambiamos las direcciones IPs a estado de pendiente 
                ip = IP.objects.get(pk=id_ip)
                ip.estado = "Pendiente"
                ip.save()
        elif accion == 'noBloquear':
            for id_ip in ids: #Cambiamos las direcciones IPs a estado de No Bloqueado 
                ip = IP.objects.get(pk=id_ip)
                ip.estado = "No Bloqueado"
                ip.save()
        elif accion == 'quitar':
            for id_ip in ids:  #Cambiamos las direcciones IPs de los casos especiales 
                Casos_Especiales.objects.get(ipEspecial=id_ip).delete()

        messages.error(request, "Acción aplicada con éxito.")
        return redirect('detector')  # Redirigir a la página principal
    else:
        messages.error(request, "Método no permitido.")
        return redirect('detector')  # Redirigir a la página principal

# Revisar para quitar el endpoint 
def endpointBloqueadas(request): # Ruta que redirige al endpoint generado de IPs Bloqueadas
    historial = Historial_IP_FW_Bloqueadas.objects.all()
    ips_bloqueadas = [item.ipBloqueada.ip for item in historial]
    data = "\n".join(ips_bloqueadas) # Tomamos todas las IPs y las enviamos como una respuesta de retorno HTML en formato de texto plano
    return HttpResponse(data, content_type='text/plain')


#=============================================================================================
#                         Funciones temporales de un solo uso 
#=============================================================================================

def cargarDatosPaises(): # Función que carga los datos de la tabla paises de un archivo externo, correr solo una vez al inicio 
    paises = pd.read_csv(os.path.join(settings.MEDIA_ROOT, 'paises.csv')) #El archivo se encuentra en la carpeta de media del proyecto

    for index, pais in paises.iterrows(): # Por cada país se crea un registro de la tabla Pais y se almacena en la DB
        nuevo = Pais(nombre=pais['nombre'] , name=pais['name'] , iso2=pais['iso2'] , iso3=pais['iso3'] , phonecod=pais['phone_code'])
        nuevo.save() 
